"""Runs the pair-level analysis and shapes the output tables."""

from __future__ import annotations

import numpy as np
import pandas as pd

from .config import Settings, load_benchmarks, load_universe
from .data import daily_returns, monthly_returns, risk_free_monthly
from .metrics import pair_stats


def build_results(prices: pd.DataFrame, settings: Settings) -> dict:
    """For every fund and every configured benchmark, compute the full stat block."""
    uni = load_universe()
    bms = load_benchmarks()
    daily = daily_returns(prices)
    monthly = monthly_returns(prices)
    rf = risk_free_monthly(prices, settings.d("rf_ticker"))

    window = int(settings.a("rolling_window_months"))
    lags = int(settings.a("newey_west_lags"))
    which = list(settings.a("benchmarks"))
    if settings.a("include_peer"):
        which = which + ["peer"]

    results: dict[tuple[str, str], dict] = {}
    skipped: list[str] = []

    for _, row in uni.iterrows():
        tkr = row["ticker"]
        if tkr not in monthly.columns or monthly[tkr].notna().sum() < 3:
            skipped.append(f"{tkr}: no return history in the loaded data")
            continue
        for kind in which:
            code = row[f"{kind}_bm"]
            if not isinstance(code, str) or not code.strip():
                continue
            if code not in monthly.columns:
                skipped.append(f"{tkr} vs {code}: benchmark series not available")
                continue
            stats = pair_stats(
                monthly[tkr], monthly[code], daily[tkr], daily[code], rf,
                window=window, nw_lags=lags,
                expense_ratio=float(row["expense_ratio"]),
            )
            if not stats:
                skipped.append(f"{tkr} vs {code}: no overlapping window")
                continue
            stats.update({
                "ticker": tkr,
                "fund_name": row["fund_name"],
                "sleeve": row["sleeve"],
                "bm_kind": kind,
                "bm_code": code,
                "bm_name": bms.loc[code, "bm_name"] if code in bms.index else code,
                "proxy_ticker": bms.loc[code, "proxy_ticker"] if code in bms.index else "",
                "proxy_quality": bms.loc[code, "proxy_quality"] if code in bms.index else "",
                "short_history": len(monthly[tkr].dropna()) < int(settings.a("min_months")),
            })
            results[(tkr, kind)] = stats

    return {"results": results, "monthly": monthly, "daily": daily, "rf": rf,
            "skipped": skipped, "universe": uni, "benchmarks": bms}


# ------------------------------------------------------------------ table shaping

_PERF_COLS = [
    ("ticker", "Ticker"), ("sleeve", "Sleeve"), ("bm_code", "Benchmark"),
    ("start_str", "Start"), ("years", "Yrs"),
    ("fund_cum", "Fund cum"), ("bench_cum", "Bm cum"),
    ("fund_ann", "Fund ann"), ("bench_ann", "Bm ann"),
    ("fund_vol", "Fund vol"), ("bench_vol", "Bm vol"),
    ("fund_sharpe", "Fund SR"), ("bench_sharpe", "Bm SR"),
    ("fund_max_dd", "Fund maxDD"), ("bench_max_dd", "Bm maxDD"),
]

_EXCESS_COLS = [
    ("ticker", "Ticker"), ("bm_code", "Benchmark"),
    ("years", "Yrs"), ("excess_ann_geom", "Excess ann"),
    ("expense_ratio", "Fee"), ("excess_before_fees", "Excess pre-fee"),
    ("tracking_error", "TE"), ("info_ratio", "IR"), ("t_stat", "t-stat"),
    ("hit_rate", "Hit rate"), ("capm_beta", "Beta"),
    ("capm_alpha_ann", "CAPM alpha"), ("capm_alpha_t", "alpha t"),
    ("up_capture", "Up cap"), ("down_capture", "Down cap"),
    ("dd_diff", "maxDD diff"),
]


def raw_frame(results: dict) -> pd.DataFrame:
    rows = []
    for (tkr, kind), st in results.items():
        row = {k: v for k, v in st.items() if not k.startswith("_")}
        rows.append(row)
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["start_str"] = pd.to_datetime(df["start"]).dt.strftime("%Y-%m")
    df["end_str"] = pd.to_datetime(df["end"]).dt.strftime("%Y-%m")
    order = {"broad": 0, "style": 1, "peer": 2}
    df = df.sort_values(["sleeve", "ticker", "bm_kind"], key=lambda s: s.map(order).fillna(s))
    return df.reset_index(drop=True)


def _fmt(df: pd.DataFrame, spec: list[tuple[str, str]]) -> pd.DataFrame:
    pct1 = {"excess_ann_geom", "expense_ratio", "excess_before_fees", "tracking_error",
            "dd_diff", "excess_ann_mean"}
    pct0 = {"fund_cum", "bench_cum", "fund_ann", "bench_ann", "fund_vol", "bench_vol",
            "fund_max_dd", "bench_max_dd", "hit_rate", "capm_alpha_ann"}
    num2 = {"fund_sharpe", "bench_sharpe", "info_ratio", "t_stat", "capm_beta",
            "capm_alpha_t", "up_capture", "down_capture", "years"}

    out = pd.DataFrame()
    for key, label in spec:
        if key not in df.columns:
            out[label] = ["n/a"] * len(df)
            continue
        col = df[key]
        if key in pct1:
            out[label] = col.map(lambda v: "n/a" if pd.isna(v) else f"{v * 100:+.2f}%")
        elif key in pct0:
            out[label] = col.map(lambda v: "n/a" if pd.isna(v) else f"{v * 100:.1f}%")
        elif key in num2:
            out[label] = col.map(lambda v: "n/a" if pd.isna(v) else f"{v:.2f}")
        else:
            out[label] = col.astype(str)
    return out


def performance_table(df: pd.DataFrame, bm_kind: str = "style") -> pd.DataFrame:
    sub = df[df["bm_kind"] == bm_kind]
    return _fmt(sub, _PERF_COLS)


def excess_table(df: pd.DataFrame, bm_kind: str = "style") -> pd.DataFrame:
    sub = df[df["bm_kind"] == bm_kind]
    return _fmt(sub, _EXCESS_COLS)


def decomposition_table(df: pd.DataFrame) -> pd.DataFrame:
    """The table that actually answers the brief: split total excess vs the broad
    benchmark into the part explained by the style tilt and the part left over."""
    broad = df[df["bm_kind"] == "broad"].set_index("ticker")
    style = df[df["bm_kind"] == "style"].set_index("ticker")
    common = [t for t in broad.index if t in style.index]
    rows = []
    for t in common:
        b, s = broad.loc[t], style.loc[t]
        vs_broad = b["excess_ann_geom"]
        vs_style = s["excess_ann_geom"]
        rows.append({
            "Ticker": t,
            "Sleeve": b["sleeve"],
            "Broad bm": b["bm_code"],
            "Style bm": s["bm_code"],
            "vs broad": vs_broad,
            "Style tilt effect": vs_broad - vs_style,
            "Implementation vs style": vs_style,
            "Impl t-stat": s["t_stat"],
            "Yrs": s["years"],
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out.sort_values("Implementation vs style", ascending=False)
    for c in ("vs broad", "Style tilt effect", "Implementation vs style"):
        out[c] = out[c].map(lambda v: "n/a" if pd.isna(v) else f"{v * 100:+.2f}%")
    out["Impl t-stat"] = out["Impl t-stat"].map(lambda v: "n/a" if pd.isna(v) else f"{v:.2f}")
    out["Yrs"] = out["Yrs"].map(lambda v: f"{v:.1f}")
    return out.reset_index(drop=True)


# ------------------------------------------------------------------ workbook export


def write_workbook(path, df: pd.DataFrame, monthly: pd.DataFrame, results: dict,
                   settings: Settings, skipped: list[str]) -> None:
    """Supporting dataset. Values only, no formulas, so nothing needs recalculating."""
    uni = load_universe(include_only=False)
    bms = load_benchmarks().reset_index()

    excess_panel = {}
    for (tkr, kind), st in results.items():
        if kind != "style":
            continue
        roll = st.get("_rolling")
        if roll is not None and not roll.empty:
            excess_panel[tkr] = roll

    notes = pd.DataFrame({
        "item": [
            "Source", "Window", "Return basis", "Frequency of statistics",
            "Excess return definition", "Significance", "Fees", "Generated",
        ],
        "detail": [
            settings.source,
            settings.a("window"),
            "market price total return, dividends reinvested" if settings.source == "yahoo"
            else "as supplied in the external files",
            "monthly for statistics, daily for drawdowns",
            "geometric difference of annualised returns, fund minus benchmark",
            "t-stat = IR * sqrt(years). Rolling window t-stats are Newey-West corrected.",
            "reported returns are already net of fund fees",
            pd.Timestamp.today().strftime("%Y-%m-%d"),
        ],
    })

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        notes.to_excel(xl, sheet_name="notes", index=False)
        uni.to_excel(xl, sheet_name="mapping_funds", index=False)
        bms.to_excel(xl, sheet_name="mapping_benchmarks", index=False)
        if not df.empty:
            df.drop(columns=[c for c in df.columns if c.startswith("_")], errors="ignore") \
              .to_excel(xl, sheet_name="stats_all", index=False)
            performance_table(df, "style").to_excel(xl, sheet_name="perf_summary", index=False)
            excess_table(df, "style").to_excel(xl, sheet_name="excess_summary", index=False)
            dec = decomposition_table(df)
            if not dec.empty:
                dec.to_excel(xl, sheet_name="decomposition", index=False)
        monthly.round(6).to_excel(xl, sheet_name="returns_monthly")
        if excess_panel:
            pd.DataFrame(excess_panel).round(6).to_excel(xl, sheet_name="rolling_excess_12m")
        if skipped:
            pd.DataFrame({"skipped": skipped}).to_excel(xl, sheet_name="exceptions", index=False)

    _style_workbook(path)


def _style_workbook(path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font

    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=9, bold=(cell.row == 1))
        ws.freeze_panes = "A2"
        for col in ws.columns:
            letter = col[0].column_letter
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 9), 46)
        ws["A1"].alignment = Alignment(horizontal="left")
    wb.save(path)
